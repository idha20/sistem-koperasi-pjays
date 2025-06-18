import io
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import render
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from student.models import Member  # Import Member model
from teacher.models import Teacher  # Import Teacher model
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from django.contrib import messages
from django.shortcuts import redirect

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def generate_report_pelajar(request):
    # Query data for students
    student = Member.objects.all()

    # Define the date style
    date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")

    if request.method == "POST":

        report_type = request.POST.get("report-type")
        tingkatan = request.POST.get("tingkatan")
        kelas = request.POST.get("kelas")
        ahli = request.POST.get("ahli")
        status = request.POST.get("status")

        # Apply filters for students based on form inputs
        if tingkatan and tingkatan != "Pilih":
            student = student.filter(tingkatan=tingkatan)
        if kelas and kelas != "Pilih":
            student = student.filter(kelas=kelas)
        if ahli and ahli != "keseluruhan":
            student = student.filter(ahli=ahli)
        if status and status != "keseluruhan":
            student = student.filter(status=status)


        # Prepare headers with filtering details
        report_headers = [
            f"Tingkatan: {tingkatan if tingkatan else 'Semua'}",
            f"Kelas: {kelas if kelas else 'Semua'}",
            f"Ahli: {ahli if ahli else 'Semua'}",
            f"Status: {status if status else 'Semua'}",
        ]

        
        if not student.exists():  # Check if there is no data
            messages.error(request, "Tiada data pelajar untuk dijana dalam laporan.")
            return redirect("generate_report_pelajar")  # Redirect back to the page

        # Handle PDF report generation
        if report_type == "pdf-type":
            buffer = io.BytesIO()

            # Set up the PDF document
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            style_normal = styles["Normal"]  # Use normal style with text wrapping

            elements = []

            # Add a title
            elements.append(Paragraph("Laporan Pelajar", styles["Title"]))
            elements.append(Spacer(1, 12))

            # Add report details
            for header in report_headers:
                elements.append(Paragraph(header, styles["Normal"]))
            elements.append(Spacer(1, 12))

            # Create the table data
            table_data = [
                ["No.", "Nama", "Tingkatan", "Kelas", "Ahli", "Status"]
            ]
            for idx, member in enumerate(student, start=1):
                table_data.append([
                    idx, 
                    Paragraph(member.nama, style_normal), 
                    Paragraph(member.tingkatan, style_normal), 
                    Paragraph(member.kelas, style_normal), 
                    Paragraph(member.ahli, style_normal),
                    Paragraph(member.status, style_normal),
                ])

            # Create and style the table
            table = Table(table_data, colWidths=[30, 200, 70, 80, 50, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Vertically center text
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),  # Font size for the table header
                ('FONTSIZE', (0, 1), (-1, -1), 8),  # Font size for the table content
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(table)

            # Add footer
            elements.append(Spacer(1, 24))  # Add some space before the footer
            elements.append(Paragraph("Disediakan Oleh: Sistem Koperasi Pjays", styles["Normal"]))

            # Build the PDF
            try:
                doc.build(elements)  # Build the PDF document
                buffer.seek(0)  # Reset buffer pointer to the beginning

                response = HttpResponse(buffer, content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="Laporan Pelajar.pdf"'
                return response

            except Exception as e:
                # Handle PDF generation errors
                return HttpResponse(f"Error generating PDF: {str(e)}")

        elif report_type == "xlsx-type":
            
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Laporan Pelajar"

            # Add table headers
            headers = [
                "No.", "Nama", "Ic Pelajar", "Jantina", "Status Pengambilan Saham", "Alamat Rumah", 
                "Tingkatan", "Kelas", "Ahli", "Modal Syer", "Tarikh Daftar"
            ]
            sheet.append(headers)

            # Styling for header row
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue color
            header_font = Font(bold=True, color="FFFFFF")  # White font
            header_alignment = Alignment(horizontal="center", vertical="center")
            border_style = Border(
                top=Side(style="thin"),
                bottom=Side(style="thin"),
                left=Side(style="thin"),
                right=Side(style="thin")
            )

            # Apply styles to header cells
            for col, value in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col)
                cell.value = value
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style

            # Add student data
            row = 2  # Start from row 2, since row 1 contains the headers
            for idx, member in enumerate(student, start=1):
                # Ensure tarikh_daftar is in date format (if it's not already)
                tarikh_daftar = member.tarikh_daftar
                if isinstance(tarikh_daftar, str):
                    tarikh_daftar = pd.to_datetime(tarikh_daftar).date()  # Convert string to date
                
                # Add the row with the student data
                sheet.append([
                    idx, member.nama, member.ic_pelajar, member.jantina, member.status,
                    member.alamat_rumah, member.tingkatan, member.kelas, member.ahli, 
                    member.modal_syer, member.tarikh_daftar.isoformat()
                ])

                # Apply the date style to the 'Tarikh Daftar' column
                date_cell = sheet.cell(row=row, column=len(headers))  # 'Tarikh Daftar' column
                date_cell.style = date_style
                row += 1

            # Apply border and alignment to all rows (header + data rows)
            total_rows = sheet.max_row  # Ensure we cover all rows dynamically
            for r in range(1, total_rows + 1):  # Include row 1 (header)
                for c in range(1, len(headers) + 1):  # Iterate through all columns
                    cell = sheet.cell(row=r, column=c)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border_style

            # Adjust column widths
            column_widths = [5, 60, 15, 10, 15, 30, 10, 15, 15, 15, 15]
            for i, width in enumerate(column_widths, start=1):
                sheet.column_dimensions[get_column_letter(i)].width = width

            # Create response to download the file
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="Laporan Pelajar.xlsx"'
            workbook.save(response)
            return response
    
        elif report_type not in ["pdf-type", "xlsx-type"]:
            messages.error(request, "Ralat yang tidak dijangka berlaku semasa memproses fail. Sila periksa format fail dan cuba lagi.")
            return redirect('generate_report_pelajar')  # Redirect to the report page instead of returning None

    # Render the report generation page
    context = {
        'student': student,
    }
    return render(request, 'generate/muka surat-Hasilkan Laporan-Pelajar.html', context)


@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def generate_report_kakitangan(request):
    # Query all kakitangan
    teachers = Teacher.objects.all()

    if request.method == "POST":

        report_type = request.POST.get("report-type")
        ahli = request.POST.get("ahli")
        status = request.POST.get("status")

        # Apply filters based on form inputs
        if ahli and ahli != "keseluruhan":
            teachers = teachers.filter(ahli=ahli)
        if status and status != "keseluruhan":
            teachers = teachers.filter(status=status)

        # Prepare headers with filtering details
        report_headers = [
            f"Ahli: {ahli if ahli else 'Semua'}",
            f"Status: {status if status else 'Semua'}",
        ]

        if not teachers.exists():  # Check if there is no data
            messages.error(request, "Tiada data kakitangan untuk dijana dalam laporan.")
            return redirect("generate_report_kakitangan")  # Redirect back to the page

        # Handle PDF report generation
        if report_type == "pdf-type":
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            style_normal = styles["Normal"]  # Use normal style with text wrapping
            elements = []

            # Add title
            elements.append(Paragraph("Laporan Kakitangan", styles["Title"]))
            elements.append(Spacer(1, 12))

            # Add report details
            for header in report_headers:
                elements.append(Paragraph(header, styles["Normal"]))
            elements.append(Spacer(1, 12))

            # Create the table data
            table_data = [
                ["No.", "Nama", "Ic", "Ahli", "Status", "Modal Syer(RM)"]
            ]
            for idx, teacher in enumerate(teachers, start=1):
                table_data.append([
                    idx,
                    Paragraph(teacher.nama, style_normal),  # Ensure this wraps
                    Paragraph(teacher.ic_cikgu, style_normal),
                    Paragraph(teacher.ahli, style_normal),
                    Paragraph(teacher.status, style_normal),
                    Paragraph(str(teacher.modal_syer), style_normal),
                ])

            # Create and style the table
            table = Table(table_data, colWidths=[30, 180, 80, 50, 80, 90])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Vertically center text
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(table)

            # Add footer
            elements.append(Spacer(1, 24))  # Add some space before the footer
            elements.append(Paragraph("Disediakan Oleh: Sistem Koperasi Pjays", styles["Normal"]))

            doc.build(elements)
            buffer.seek(0)

            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="Laporan Kakitangan.pdf"'
            return response

        # Handle Excel report generation
        elif report_type == "xlsx-type":

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Laporan Kakitangan"

            # Add table headers
            headers = ["No.", "Nama", "No. IC", "Jantina", "Alamat Rumah","Pangkat", "Ahli", "Status", "Modal Syer (RM)", "Tarikh Daftar"]

            sheet.append(headers)

            # Styling for header row
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue color
            header_font = Font(bold=True, color="FFFFFF")  # White font
            header_alignment = Alignment(horizontal="center", vertical="center")
            border_style = Border(
                top=Side(style="thin"),
                bottom=Side(style="thin"),
                left=Side(style="thin"),
                right=Side(style="thin")
            )

            # Apply styles to header cells
            for col, value in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col)
                cell.value = value
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style

                        # Add teacher data
            for idx, teacher in enumerate(teachers, start=1):
                sheet.append([
                    idx, teacher.nama, teacher.ic_cikgu, teacher.jantina, teacher.alamat_rumah,
                    teacher.pangkat, teacher.ahli, teacher.status, teacher.modal_syer, teacher.tarikh_daftar.isoformat()
                ])

            # Apply border and alignment to all rows (header + data rows)
            total_rows = sheet.max_row  # Ensure we cover all rows dynamically
            for r in range(1, total_rows + 1):  # Include row 1 (header)
                for c in range(1, len(headers) + 1):  # Iterate through all columns
                    cell = sheet.cell(row=r, column=c)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border_style

            # Add footer
            footer_row = sheet.max_row + 2  # Leave a blank row after the data
            sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(headers))
            footer_cell = sheet.cell(row=footer_row, column=1)
            footer_cell.value = "Disediakan Oleh: Sistem Koperasi Pjays"
            footer_cell.font = Font(italic=True, size=10)  # Italic font with a smaller size
            footer_cell.alignment = Alignment(horizontal="right", vertical="center")  # Align to the right


            # Adjust column widths
            column_widths = [5, 60, 20, 15, 40, 25, 10, 15, 15, 15]
            for i, width in enumerate(column_widths, start=1):
                sheet.column_dimensions[chr(64 + i)].width = width

            # Create response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="Laporan Kakitangan.xlsx"'
            workbook.save(response)
            return response

        else:
            messages.error(request, "Ralat yang tidak dijangka berlaku semasa memproses fail. Sila periksa format fail dan cuba lagi.")
            return redirect('generate_report_kakitangan')  # Redirect to the report page instead of returning None

    # Render the report generation page
    context = {
        'teacher': teachers,
    }
    return render(request, 'generate/muka surat-Hasilkan Laporan-Kakitangan.html', context)

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def generate_report_saham(request):
    student = Member.objects.all()  # Query for students
    teachers = Teacher.objects.all()  # Query for teachers

    # Define the date style
    date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")

    if request.method == "POST":
        komuniti = request.POST.get("komuniti")

        if komuniti == "pelajar":
            
            report_type = request.POST.get("report-type")
            tingkatan = request.POST.get("tingkatan")
            kelas = request.POST.get("kelas")
            ahli = request.POST.get("ahli")
            status = request.POST.get("status")

            # Apply filters for students based on form inputs
            if tingkatan and tingkatan != "Pilih":
                student = student.filter(tingkatan=tingkatan)
            if kelas and kelas != "Pilih":
                student = student.filter(kelas=kelas)
            if ahli and ahli != "keseluruhan":
                student = student.filter(ahli=ahli)
            if status and status != "keseluruhan":
                student = student.filter(status=status)


            # Prepare headers with filtering details
            report_headers = [
                f"Tingkatan: {tingkatan if tingkatan else 'Semua'}",
                f"Kelas: {kelas if kelas else 'Semua'}",
                f"Ahli: {ahli if ahli else 'Semua'}",
                f"Status: {status if status else 'Semua'}",
            ]

            if not student.exists():  # Check if there is no data
                messages.error(request, "Tiada data pelajar untuk dijana dalam laporan.")
                return redirect("generate_report_saham")  # Redirect back to the page

            # Handle PDF report generation
            if report_type == "pdf-type":
            
                buffer = io.BytesIO()

                # Set up the PDF document
                doc = SimpleDocTemplate(buffer, pagesize=letter)
                styles = getSampleStyleSheet()
                style_normal = styles["Normal"]  # Use normal style with text wrapping

                elements = []

                # Add a title
                elements.append(Paragraph("Laporan Saham Pelajar", styles["Title"]))
                elements.append(Spacer(1, 12))

                # Add report details
                for header in report_headers:
                    elements.append(Paragraph(header, styles["Normal"]))
                elements.append(Spacer(1, 12))

                # Create the table data
                table_data = [
                    ["No.", "Nama", "Tingkatan", "Kelas", "Modal Syer","Tandatangan"]
                ]
                for idx, member in enumerate(student, start=1):
                    table_data.append([
                        idx, 
                        Paragraph(member.nama, style_normal), 
                        Paragraph(member.tingkatan, style_normal), 
                        Paragraph(member.kelas, style_normal), 
                        Paragraph(str(member.modal_syer), style_normal),
                        Paragraph(" ", style_normal),
                    ])

                # Create and style the table
                table = Table(table_data, colWidths=[30, 200, 70, 80, 70, 80])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Vertically center text
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),  # Font size for the table header
                    ('FONTSIZE', (0, 1), (-1, -1), 8),  # Font size for the table content
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))

                elements.append(table)

                # Add footer
                elements.append(Spacer(1, 24))  # Add some space before the footer
                elements.append(Paragraph("Disediakan Oleh: Sistem Koperasi Pjays", styles["Normal"]))

                # Build the PDF
                try:
                    doc.build(elements)  # Build the PDF document
                    buffer.seek(0)  # Reset buffer pointer to the beginning

                    response = HttpResponse(buffer, content_type='application/pdf')
                    response['Content-Disposition'] = 'attachment; filename="Laporan Saham Pelajar.pdf"'
                    return response

                except Exception as e:
                    # Handle PDF generation errors
                    return HttpResponse(f"Error generating PDF: {str(e)}")

            elif report_type == "xlsx-type":
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Laporan Saham Pelajar"

                # Add table headers
                headers = [
                    "No.", "Nama", "Ic Pelajar", "Tingkatan", "Kelas", "Modal Syer", "Tarikh Daftar", "Tandatangan"
                ]
                sheet.append(headers)

                # Styling for header row
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue color
                header_font = Font(bold=True, color="FFFFFF")  # White font
                header_alignment = Alignment(horizontal="center", vertical="center")
                border_style = Border(
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                    left=Side(style="thin"),
                    right=Side(style="thin")
                )

                # Apply styles to header cells
                for col, value in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col)
                    cell.value = value
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border_style

                # Add student data
                row = 2  # Start from row 2, since row 1 contains the headers
                for idx, member in enumerate(student, start=1):
                    # Ensure tarikh_daftar is in date format (if it's not already)
                    tarikh_daftar = member.tarikh_daftar
                    if isinstance(tarikh_daftar, str):
                        tarikh_daftar = pd.to_datetime(tarikh_daftar).date()  # Convert string to date
                    
                    # Add the row with the student data
                    sheet.append([
                        idx, member.nama, member.ic_pelajar, member.tingkatan, member.kelas,
                        member.modal_syer, member.tarikh_daftar.isoformat()
                    ])

                    # Apply the date style to the 'Tarikh Daftar' column
                    date_cell = sheet.cell(row=row, column=len(headers))  # 'Tarikh Daftar' column
                    date_cell.style = date_style
                    row += 1

                # Apply border and alignment to all rows (header + data rows)
                total_rows = sheet.max_row  # Ensure we cover all rows dynamically
                for r in range(1, total_rows + 1):  # Include row 1 (header)
                    for c in range(1, len(headers) + 1):  # Iterate through all columns
                        cell = sheet.cell(row=r, column=c)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border_style

                # Add footer
                footer_row = sheet.max_row + 2  # Leave a blank row after the data
                sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(headers))
                footer_cell = sheet.cell(row=footer_row, column=1)
                footer_cell.value = "Disediakan Oleh: Sistem Koperasi Pjays"
                footer_cell.font = Font(italic=True, size=10)  # Italic font with a smaller size
                footer_cell.alignment = Alignment(horizontal="right", vertical="center")  # Align to the right

                # Adjust column widths
                column_widths = [5, 60, 15, 10, 15, 20, 20, 25]
                for i, width in enumerate(column_widths, start=1):
                    sheet.column_dimensions[get_column_letter(i)].width = width

                # Create response to download the file
                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = 'attachment; filename="Laporan Saham Pelajar.xlsx"'
                workbook.save(response)
                return response
        
            elif report_type not in ["pdf-type", "xlsx-type"]:
                messages.error(request, "Ralat yang tidak dijangka berlaku semasa memproses fail. Sila periksa format fail dan cuba lagi.")
                return redirect('generate_report_saham')  # Redirect to the report page instead of returning None
            
        if komuniti == "kakitangan":

            report_type = request.POST.get("report-type")
            ahli = request.POST.get("ahli_kakitangan")
            status = request.POST.get("status_kakitangan")

            # Apply filters based on form inputs
            if ahli and ahli != "keseluruhan":
                teachers = teachers.filter(ahli=ahli)
            if status and status != "keseluruhan":
                teachers = teachers.filter(status=status)

            # Prepare headers with filtering details
            report_headers = [
                f"Ahli: {ahli if ahli else 'Semua'}",
                f"Status: {status if status else 'Semua'}",
            ]

            if not teachers.exists():  # Check if there is no data
                messages.error(request, "Tiada data kakitangan untuk dijana dalam laporan.")
                return redirect("generate_report_saham")  # Redirect back to the page

            # Handle PDF report generation
            if report_type == "pdf-type":

                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)
                styles = getSampleStyleSheet()
                style_normal = styles["Normal"]  # Use normal style with text wrapping
                elements = []

                # Add title
                elements.append(Paragraph("Laporan Saham Kakitangan", styles["Title"]))
                elements.append(Spacer(1, 12))

                # Add report details
                for header in report_headers:
                    elements.append(Paragraph(header, styles["Normal"]))
                elements.append(Spacer(1, 12))

                # Create the table data
                table_data = [
                    ["No.", "Nama", "Ic", "Ahli", "Status", "Modal Syer(RM)", "Tandatangan"]
                ]
                for idx, teacher in enumerate(teachers, start=1):
                    table_data.append([
                        idx,
                        Paragraph(teacher.nama, style_normal),  # Ensure this wraps
                        Paragraph(teacher.ic_cikgu, style_normal),
                        Paragraph(teacher.ahli, style_normal),
                        Paragraph(teacher.status, style_normal),
                        Paragraph(str(teacher.modal_syer), style_normal),
                        Paragraph(" ", style_normal),
                    ])

                # Create and style the table
                table = Table(table_data, colWidths=[30, 180, 80, 50, 60, 90, 70])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Vertically center text
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))

                elements.append(table)

                # Add footer
                elements.append(Spacer(1, 24))  # Add some space before the footer
                elements.append(Paragraph("Disediakan Oleh: Sistem Koperasi Pjays", styles["Normal"]))

                doc.build(elements)
                buffer.seek(0)

                response = HttpResponse(buffer, content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="Laporan Saham Kakitangan.pdf"'
                return response

            # Handle Excel report generation
            elif report_type == "xlsx-type":

                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Laporan Saham Kakitangan"

                # Add table headers
                headers = ["No.", "Nama", "No. IC", "Ahli", "Status", "Modal Syer (RM)", "Tarikh Daftar", "Tandatangan"]

                sheet.append(headers)

                # Styling for header row
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue color
                header_font = Font(bold=True, color="FFFFFF")  # White font
                header_alignment = Alignment(horizontal="center", vertical="center")
                border_style = Border(
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                    left=Side(style="thin"),
                    right=Side(style="thin")
                )

                # Apply styles to header cells
                for col, value in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col)
                    cell.value = value
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border_style

                            # Add teacher data
                for idx, teacher in enumerate(teachers, start=1):
                    sheet.append([
                        idx, teacher.nama, teacher.ic_cikgu, teacher.ahli, teacher.status, 
                        teacher.modal_syer, teacher.tarikh_daftar.isoformat()
                    ])

                # Apply border and alignment to all rows (header + data rows)
                total_rows = sheet.max_row  # Ensure we cover all rows dynamically
                for r in range(1, total_rows + 1):  # Include row 1 (header)
                    for c in range(1, len(headers) + 1):  # Iterate through all columns
                        cell = sheet.cell(row=r, column=c)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border_style

                # Add footer
                footer_row = sheet.max_row + 2  # Leave a blank row after the data
                sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(headers))
                footer_cell = sheet.cell(row=footer_row, column=1)
                footer_cell.value = "Disediakan Oleh: Sistem Koperasi Pjays"
                footer_cell.font = Font(italic=True, size=10)  # Italic font with a smaller size
                footer_cell.alignment = Alignment(horizontal="right", vertical="center")  # Align to the right


                # Adjust column widths
                column_widths = [5, 60, 20, 25, 15, 15, 25, 15]
                for i, width in enumerate(column_widths, start=1):
                    sheet.column_dimensions[chr(64 + i)].width = width

                # Create response
                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = 'attachment; filename="Laporan Saham Kakitangan.xlsx"'
                workbook.save(response)
                return response
            
            else:
                messages.error(request, "Ralat yang tidak dijangka berlaku semasa memproses fail. Sila periksa format fail dan cuba lagi.")
                return redirect('generate_report_saham')  # Redirect to the report page instead of returning None

    context = {
        'student': student,
        'teacher': teachers
    }

    return render(request, 'generate/muka surat-Hasilkan Laporan-Saham.html', context)

@login_required(login_url='login')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def generate_report_keseluruhan(request):
    student = Member.objects.all()  # Query for students
    teacher = Teacher.objects.all()  # Query for teachers

    print("Teachers:", teacher)  # Print to check if data is being returned

    context = {
        'student': student,
        'teacher': teacher
    }

    return render(request, 'generate/muka surat-Hasilkan Laporan-Keseluruhan.html', context)

